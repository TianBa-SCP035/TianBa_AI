# -*- coding: utf-8 -*-
"""表3:受试品对小鼠肿瘤重量抑瘤作用"""
import re
import sys
import json
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from graphpad_helper import add_graphpad_sheet

# ========== 配置（精简但不简化业务） ==========
CFG = {
    "SHEET_DATA": ["样品收集方案", "Sample Collection Record"],
    "SHEET_DESIGN": ["实验设计", "Study Design"],
    "OUT_SHEET": "form_7_3",

    # 仅保留中文 + 一个英文关键词
    "GROUP_PATTERN": r"^\s*G\d+\b",
    "TUMOR_HEADERS": ["肿瘤", "Tumor"],         # 用于定位"肿瘤重量"所在列的表头单元格
    "MEAN_LABELS": ["均数", "Average"],
    "SD_LABELS": ["标准误", "Standard Error of the Mean"],
    "TGITW_LABELS": ["TGITW"],       # TGITW 行标签
    "DESIGN_GROUP_HEADER": ["组别", "Groups"],
    "DESIGN_DRUG_HEADERS": ["处理方式", "Treatment"],
    "DESIGN_DOSE_HEADERS": ["剂量", "Dosages"],

    "CONTROL_GROUP": "G1",
    "DECIMALS_TW": 3,     # 瘤重均数±SD 小数位
    "DECIMALS_TGI": 1,    # TGITW(%) 小数位
}

# ========== 小工具（精简实现） ==========
def norm(s) -> str:
    if s is None:
        return ""
    s = str(s).replace("\u3000", " ").strip()
    return re.sub(r"\s+", " ", s)

def contains_any(text, patterns) -> bool:
    """检查文本是否包含任意一个模式字符串；patterns 可为 str 或 list"""
    if not text or not patterns:
        return False
    if isinstance(patterns, str):
        patterns = [patterns]
    # 对文本和关键词都进行标准化并移除括号和空格
    text_norm = re.sub(r"[（）()\s]", "", norm(text))
    for pattern in patterns:
        pattern_norm = re.sub(r"[（）()\s]", "", norm(pattern))
        if pattern_norm in text_norm:
            return True
    return False

def in_merged(ws: Worksheet, r: int, c: int):
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:
            return rng
    return None

def val_eff(ws: Worksheet, r: int, c: int):
    v = ws.cell(row=r, column=c).value
    if v is not None:
        return v
    rng = in_merged(ws, r, c)
    if rng:
        return ws.cell(rng.min_row, rng.min_col).value
    return None

def is_empty(ws: Worksheet, r: int, c: int) -> bool:
    return norm(val_eff(ws, r, c)) == ""

def parse_float(x):
    s = norm(x)
    if s in ("", "-", "—", "–"):
        return None
    s = s.replace(",", "").replace("%", "")
    try:
        return float(s)
    except Exception:
        m = re.search(r"(-?\d+(?:\.\d+)?)", s)
        return float(m.group(1)) if m else None

def fmt_pm(m, sd, d=1):
    if m is None or sd is None:
        return ""
    # 添加容差处理，确保与Excel四舍五入一致
    return f"{round(m + 1e-06, d):.{d}f}±{round(sd + 1e-06, d):.{d}f}"

def auto_adjust_column_width(ws):
    """自动调整工作表列宽以适应内容"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 6, 50)  # 限制最大宽度为50
        ws.column_dimensions[column_letter].width = adjusted_width

def find_existing_sheet(wb, sheet_names) -> str:
    """从候选工作表名称中返回第一个存在的名称；否则返回空字符串"""
    if isinstance(sheet_names, str):
        sheet_names = [sheet_names]
    for name in sheet_names or []:
        if name in wb.sheetnames:
            return name
    return ""

# ========== 主流程 ==========
def extract_table(xlsx_in: str, xlsx_out: str) -> bool:
    from P_compute import calculate_dunnett_json  # 复用你原来的 Dunnett 计算

    C = CFG
    try:
        wb = load_workbook(xlsx_in, data_only=True)

        # 1) 数据页：样品收集方案
        sheet_name = find_existing_sheet(wb, C["SHEET_DATA"])
        if not sheet_name:
            raise RuntimeError(f"未找到数据页：{C['SHEET_DATA']}")
        ws = wb[sheet_name]

        # 2) 定位"肿瘤/Tumor"表头列 → 数据列 = 左一列（左列不数值则回退本列）
        tumor_col = None
        for r in range(1, min(50, ws.max_row) + 1):
            for c in range(1, min(50, ws.max_column) + 1):
                t = val_eff(ws, r, c)
                if contains_any(t, C["TUMOR_HEADERS"]):
                    tumor_col = c
                    break
            if tumor_col: break
        if not tumor_col:
            raise RuntimeError("未找到包含\"肿瘤/Tumor\"的列头。")

        data_col = tumor_col - 1 if tumor_col > 1 else tumor_col

        def seems_numeric_col(col, sample_rows=40):
            cnt = 0
            for rr in range(1, min(ws.max_row, sample_rows) + 1):
                if parse_float(val_eff(ws, rr, col)) is not None:
                    cnt += 1
            return cnt >= 3

        if not seems_numeric_col(data_col) and seems_numeric_col(tumor_col):
            data_col = tumor_col

        # 3) 组起点：A列匹配 G\d+
        patG = re.compile(C["GROUP_PATTERN"], re.IGNORECASE)
        group_starts = [r for r in range(1, ws.max_row + 1) if patG.match(norm(val_eff(ws, r, 1)))]
        if not group_starts:
            raise RuntimeError("未在 A 列识别到任何组别（G1/G2/...）。")

        rows_out = []
        long_rows = []  # 用于 Dunnett
        per_group_values = {}  # 用于GraphPad工作表
        # 4) 逐组解析：B列识别统计行，数据列读取
        for i, rs in enumerate(group_starts):
            re_ = group_starts[i + 1] - 1 if i < len(group_starts) - 1 else ws.max_row
            gname = norm(val_eff(ws, rs, 1))

            def find_row(labels):
                for rr in range(rs, re_ + 1):
                    t = norm(val_eff(ws, rr, 2))  # B列
                    if any(lbl == t or lbl in t for lbl in labels):
                        return rr
                return None

            r_mean = find_row(C["MEAN_LABELS"])
            r_sd   = find_row(C["SD_LABELS"])
            r_tgi  = find_row(C["TGITW_LABELS"])

            m  = parse_float(val_eff(ws, r_mean, data_col)) if r_mean else None
            sd = parse_float(val_eff(ws, r_sd,   data_col)) if r_sd   else None
            tw_fmt = fmt_pm(m, sd, C["DECIMALS_TW"])

            tgi = None
            if r_tgi:
                raw = val_eff(ws, r_tgi, data_col)
                vn = parse_float(raw)
                if vn is not None:
                    tgi = vn * 100 if (vn <= 1 and "%" not in str(raw)) else vn
            tgi_fmt = "-" if (tgi is None or gname == C["CONTROL_GROUP"]) \
                      else f"{round(tgi + 1e-06, C['DECIMALS_TGI']):.{C['DECIMALS_TGI']}f}"

            # 个体原始值：组别行 ~ 均数行-1
            end_anim = (r_mean - 1) if r_mean else re_
            values = []  # 用于GraphPad工作表
            for rr in range(rs, end_anim + 1):
                v = parse_float(val_eff(ws, rr, data_col))
                if v is not None:
                    long_rows.append({"group": gname, "volume": float(v)})
                    values.append(v)
            per_group_values[gname] = values

            rows_out.append({
                "组别": gname,
                "瘤重": tw_fmt,
                "TGITW": tgi_fmt,
                # P值稍后统一回填
            })

        df = pd.DataFrame(rows_out)

        # 5) 受试品：从"实验设计/Study Design"页简洁映射 组别→处理方式(剂量)
        sheet_design_name = find_existing_sheet(wb, C["SHEET_DESIGN"])
        if sheet_design_name:
            wsD = wb[sheet_design_name]
            # 找包含三列名的表头行（中文或英文）
            hdr = None
            for r in range(1, wsD.max_row + 1):
                row_txt = " ".join(norm(val_eff(wsD, r, c)) for c in range(1, wsD.max_column + 1))
                if contains_any(row_txt, C["DESIGN_GROUP_HEADER"]) and \
                   contains_any(row_txt, C["DESIGN_DRUG_HEADERS"]) and \
                   contains_any(row_txt, C["DESIGN_DOSE_HEADERS"]):
                    hdr = r
                    break
            if hdr is None:
                raise RuntimeError("实验设计页未找到表头。")
                
            col_g = col_d = col_do = None
            for c in range(1, wsD.max_column + 1):
                t = norm(val_eff(wsD, hdr, c))
                if (col_g is None) and contains_any(t, C["DESIGN_GROUP_HEADER"]):
                    col_g = c
                if (col_d is None) and contains_any(t, C["DESIGN_DRUG_HEADERS"]):
                    col_d = c
                if (col_do is None) and contains_any(t, C["DESIGN_DOSE_HEADERS"]):
                    col_do = c

            mapping = {}
            blank = 0
            for r in range(hdr + 1, wsD.max_row + 1):
                g = norm(val_eff(wsD, r, col_g)) if col_g else ""
                if g == "":
                    blank += 1
                    if blank >= 2:
                        break
                    continue
                blank = 0
                drug = norm(val_eff(wsD, r, col_d)) if col_d else ""
                dose = norm(val_eff(wsD, r, col_do)) if col_do else ""
                combo = f"{drug}({dose})" if (drug and dose) else (drug or (f"({dose})" if dose else ""))
                if combo:
                    mapping.setdefault(g, []).append(combo)

            if mapping:
                df_map = pd.DataFrame([{"组别": k, "受试品": ", ".join(v)} for k, v in mapping.items()])
                df = df.merge(df_map, on="组别", how="left")
                df.insert(1, "受试品", df.pop("受试品"))
            else:
                df.insert(1, "受试品", "")

        # 6) Dunnett 计算 P 值（基于 long_rows 的个体值；对照组留空）
        p_map = {}
        if long_rows:
            res = calculate_dunnett_json(json.dumps(long_rows, ensure_ascii=False),
                                         control=C["CONTROL_GROUP"])
            for r in res:
                g = (r.get("group") or "").strip()
                stars = (r.get("Summary") or "").strip()   # '**' / '' / 'ns'
                pval  = (r.get("P-Value") or "").strip()   # '0.0056' 等
                p_map[g] = "" if g == C["CONTROL_GROUP"] else (pval if stars == "ns" else f"{stars}{pval}")

        df["P值"] = df["组别"].map(lambda g: p_map.get(g, "" if g == C["CONTROL_GROUP"] else "-"))

        # 7) 列顺 & 输出
        df = df[["组别", "受试品", "瘤重", "TGITW", "P值"]].fillna("-").replace("", "-")

        # 写出
        import os
        if not os.path.exists(xlsx_out):
            book = Workbook()
            default = book.active
            book.remove(default)
            book.save(xlsx_out)

        book = load_workbook(xlsx_out)
        if C["OUT_SHEET"] in book.sheetnames:
            book.remove(book[C["OUT_SHEET"]])
        ws_out = book.create_sheet(C["OUT_SHEET"])

        # 表头 + 数据
        for j, col in enumerate(df.columns, 1):
            ws_out.cell(1, j, col)
        for i, row in enumerate(df.itertuples(index=False), start=2):
            for j, v in enumerate(row, start=1):
                ws_out.cell(i, j, v)

        auto_adjust_column_width(ws_out)
        book.save(xlsx_out)
        
        # 添加GraphPad工作表
        if per_group_values:
            success = add_graphpad_sheet(xlsx_out, "7-3实验动物瘤重数据", per_group_values)
            if success:
                print(f"成功添加GraphPad使用工作表")
            else:
                print(f"添加GraphPad使用工作表失败")
        
        print(f"OK: {xlsx_out} / {C['OUT_SHEET']}  共 {len(df)} 行")
        return True

    except Exception as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        return False


if __name__ == "__main__":
    input_path = "D:/TianBa_AI/Code/docs/temp/project_report/25P123501_Final.xlsx"
    output_path = "D:/TianBa_AI/Code/docs/temp/project_report/25P123501_Detail.xlsx"
    
    ok = extract_table(input_path, output_path)
    sys.exit(0 if ok else 1)
