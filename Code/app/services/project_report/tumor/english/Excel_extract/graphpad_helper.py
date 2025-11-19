# -*- coding: utf-8 -*-
"""通用辅助函数:添加GraphPad使用工作表"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def add_graphpad_sheet(workbook_path, sheet_title, per_group_values):
    """在Excel文件中添加或更新"GraphPad使用"工作表"""
    try:
        book = load_workbook(workbook_path)
        graphpad_sheet_name = "GraphPad使用"
        
        if graphpad_sheet_name in book.sheetnames:
            ws_graphpad = book[graphpad_sheet_name]
            # 找最后一行有数据的位置
            last_row = max((cell.row for row in ws_graphpad.iter_rows() 
                           for cell in row if cell.value and str(cell.value).strip()), default=0)
            start_row = last_row + 2  # 空一行
        else:
            ws_graphpad = book.create_sheet(graphpad_sheet_name)
            start_row = 1
        
        # 写入标题
        ws_graphpad.cell(row=start_row, column=1, value=sheet_title)
        
        # 写入组别表头
        group_names = list(per_group_values.keys())
        for col_idx, group_name in enumerate(group_names, 1):
            ws_graphpad.cell(row=start_row + 1, column=col_idx, value=group_name)
        
        # 按列写入每组数据
        for col_idx, group_name in enumerate(group_names, 1):
            values = per_group_values.get(group_name, [])
            for row_idx, value in enumerate(values, start_row + 2):
                ws_graphpad.cell(row=row_idx, column=col_idx, value=value)
        
        auto_adjust_column_width(ws_graphpad)
        book.save(workbook_path)
        return True
        
    except Exception as e:
        print(f"[ERROR] 添加GraphPad工作表失败: {e}")
        return False

def auto_adjust_column_width(ws):
    """自动调整列宽"""
    for column in ws.columns:
        max_length = max((len(str(cell.value)) for cell in column if cell.value), default=0)
        column_letter = get_column_letter(column[0].column)
        adjusted_width = min(max_length + 3, 30)
        ws.column_dimensions[column_letter].width = adjusted_width