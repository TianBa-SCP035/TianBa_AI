from openpyxl import load_workbook
from openpyxl.styles import Alignment
from decimal import Decimal, InvalidOperation
import re

# 遍历所有sheet名,从中提取最大的结束天数
def extract_max_end_day(excel_file: str) -> int:
    try:
        wb = load_workbook(excel_file, data_only=True, read_only=True)
        max_day = 0  # 默认值改为0，确保只有实际找到的天数才会被返回
        for sheet_name in wb.sheetnames:
            match = re.search(r'分组后第(\d+)天', sheet_name)
            if match:
                day = int(match.group(1))
                max_day = max(max_day, day)
        wb.close()
        return max_day
    except Exception as e:
        print(f"提取结束天数时发生异常: {e}")
        return 0  # 发生异常时返回0

#格式化数字
def as_text(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    try:
        d = Decimal(s)
        if d == d.to_integral_value():
            return str(d.quantize(Decimal('1')))
        return format(d, 'f').rstrip('0').rstrip('.')
    except (InvalidOperation, ValueError):
        return s

# 使用正则表达式匹配日期格式并转换格式
def convert_date_format(date_str: str) -> str:
    if not date_str:
        return date_str
    match = re.match(r'(\d{4})-(\d{1,2})-(\d{1,2})\s+\d{2}:\d{2}:\d{2}', str(date_str))
    if match:
        year, month, day = match.groups()
        return f"{year}年{month.zfill(2)}月{day.zfill(2)}日"
    return date_str

#从实验动物品系值中删除"mice"
def remove_mice_from_strain(strain_value: str) -> str:
    if not strain_value:
        return strain_value
    strain_str = str(strain_value).strip()
    if "mice" in strain_str.lower():
        result = re.sub(r'\s*mice\s*', '', strain_str, flags=re.IGNORECASE)
        return result.strip()
    return strain_str

def update_supplement_info(src_file: str, dst_file: str, user_end_day: int = None) -> int:
    """更新补充信息页，写入实验终点天和结束天；返回：实际使用的结束天数"""
    try:
        src_sheet = "项目操作信息"
        dst_sheet = "明细"
        # 读取源数据
        src_wb = load_workbook(src_file, data_only=True)
        src_ws = src_wb[src_sheet]
        # 找"实验类型"起始行
        start_row = None
        for r in range(1, src_ws.max_row + 1):
            if src_ws.cell(r, 1).value == "实验类型":
                start_row = r
                break
        if start_row is None:
            print("未找到起始行'实验类型'")
            return None

        # 收集 key->value，读取时就转成文本，切断科学计数法
        new_data = {}
        for r in range(start_row, src_ws.max_row + 1):
            k = src_ws.cell(r, 1).value
            v = src_ws.cell(r, 2).value
            if k is None and v is None:
                break
            if k is not None:
                new_data[str(k).strip()] = as_text(v)

        # 获取实验终点天（从Excel提取的最大天数）
        experiment_end_day = extract_max_end_day(src_file)
        new_data["实验终点天"] = str(experiment_end_day)
        # 如果用户提供了有效的结束天（不为None和0），则使用用户提供的值
        end_day = user_end_day if user_end_day not in [None, 0] else experiment_end_day
        new_data["结束天"] = str(end_day)
        src_wb.close()

        # 更新目标文件
        wb = load_workbook(dst_file)
        ws = wb[dst_sheet]

        # 创建字段到行号的索引
        field_row = {
            str(ws.cell(r, 1).value).strip(): r
            for r in range(1, ws.max_row + 1)
            if ws.cell(r, 1).value }

        # 更新数据
        for k, v in new_data.items():
            if k in field_row:
                row = field_row[k]
            else:
                ws.append([k, None])
                row = ws.max_row
            
            cell = ws.cell(row, 2)
            cell.number_format = '@'
            cell.data_type = 's'
            cell.value = v
            cell.alignment = Alignment(horizontal='left')

        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        
        # 处理日期格式
        for r in range(1, ws.max_row + 1):
            cell_value = ws.cell(r, 2).value
            if cell_value:
                converted_value = convert_date_format(cell_value)
                if converted_value != cell_value:
                    ws.cell(r, 2).value = converted_value
        
        # 处理实验动物品系
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value and str(ws.cell(r, 1).value).strip() == "实验动物品系":
                second_col_value = ws.cell(r, 2).value
                if second_col_value:
                    processed_value = remove_mice_from_strain(second_col_value)
                    if processed_value != second_col_value:
                        ws.cell(r, 2).value = processed_value
        
        wb.save(dst_file)
        return end_day
    except Exception as e:
        print(f"更新失败: {e}")
        return None

if __name__ == "__main__":
    # 使用完整文件路径
    src_file = "D:/TianBa_AI/Code/docs/temp/project_report/25P080002_Final.xlsx"
    dst_file = "D:/TianBa_AI/Code/docs/temp/project_report/25P080002_明细.xlsx"
    result = update_supplement_info(src_file, dst_file,10)
    print(f"文件更新成功！实际使用的结束天: {result}" if result is not None else "文件更新失败！")
