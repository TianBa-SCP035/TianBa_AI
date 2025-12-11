# -*- coding: utf-8 -*-
"""表2:实验动物荷瘤体积 mm3"""
import re
import sys
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from graphpad_helper import add_graphpad_sheet

# =============== 变量区（所有可调参数都在这里） ===============
CONFIG = {
    # 工作表与输出
    "SHEET_DATA": ["实验数据汇总", "Study Data"],     # 荷瘤体积所在页
    "SHEET_DESIGN": ["实验设计", "Study Design"],     # 实验设计页（可不存在）
    "OUT_SHEET": "form_7_2",                         # 输出工作表名称

    # 关键文本/模式（中文中可能出现全角空格或括号，后续会正则规整）
    "ANCHOR_CONTAINS": ["实验动物荷瘤体积", "Animal Tumor Volume (mm3)"],
    "DAYS_HEADER": ["分组后天数", "Days Post Grouping"],

    # 统计行识别（位于组别块内第二列）
    "MEAN_LABELS": ["均数", "Average"],
    "SD_LABELS": ["标准误", "Standard Error of the Mean"],
    "TGITV_LABELS": ["TGITV"],

    # 设计页表头（允许出现任意一个候选）
    "DESIGN_GROUP_HEADER": ["组别", "Groups"],
    "DESIGN_DRUG_HEADERS": ["处理方式", "Treatment"],
    "DESIGN_DOSE_HEADERS": ["剂量", "Dosages"],

    # 其它启发式
    "GROUP_PATTERN": r"^\s*G\d+\b",   # 组别识别（首列形如 G1/G2/...）
    "DAYS_ROW_LOOKAHEAD": 5,          # 在锚点下方多少行窗口内判定"结束列"
    "DECIMALS": 1,                    # 均值/差值/TGITV 保留小数位
}

# =============== 基础工具 ===============
def norm(s) -> str:
    if s is None:
        return ""
    s = str(s).replace("\u3000", " ").strip()
    return re.sub(r"\s+", " ", s)

def contains_any(text, patterns) -> bool:
    """检查文本是否包含任意一个模式字符串；patterns 可为 str 或 list"""
    if not text or not patterns:
        return False
    if isinstance(patterns, str):
        patterns = [patterns]
    # 对文本和关键词都进行标准化并移除括号和空格
    text_norm = re.sub(r"[（）()\s]", "", norm(text))
    for pattern in patterns:
        pattern_norm = re.sub(r"[（）()\s]", "", norm(pattern))
        if pattern_norm in text_norm:
            return True
    return False

def find_existing_sheet(wb, sheet_names) -> str:
    """从候选工作表名称中返回第一个存在的名称；否则返回空字符串"""
    if isinstance(sheet_names, str):
        sheet_names = [sheet_names]
    for name in sheet_names or []:
        if name in wb.sheetnames:
            return name
    return ""

def in_merged(ws: Worksheet, r: int, c: int):
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:
            return rng
    return None

def val_eff(ws: Worksheet, r: int, c: int):
    v = ws.cell(row=r, column=c).value
    if v is not None:
        return v
    rng = in_merged(ws, r, c)
    if rng:
        return ws.cell(rng.min_row, rng.min_col).value
    return None

def is_empty(ws: Worksheet, r: int, c: int) -> bool:
    return norm(val_eff(ws, r, c)) == ""

def parse_float(x):
    s = norm(x)
    if s in ("", "-", "—", "–"):
        return None
    s = s.replace(",", "").replace("%", "")
    try:
        return float(s)
    except Exception:
        m = re.search(r"(-?\d+(?:\.\d+)?)", s)
        return float(m.group(1)) if m else None

def fmt_pm(m, sd, d=1):
    if m is None or sd is None:
        return ""
    try:
        if d == 0:  # 如果d=0，表示只保留整数
            # 添加容差处理，确保与Excel四舍五入一致
            return f"{int(round(float(m) + 1e-06))}±{int(round(float(sd) + 1e-06))}"
        # 添加容差处理，确保与Excel四舍五入一致
        return f"{round(float(m) + 1e-06, d):.{d}f}±{round(float(sd) + 1e-06, d):.{d}f}"
    except (ValueError, TypeError):
        return ""

def auto_adjust_column_width(ws):
    """自动调整工作表列宽以适应内容"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 6, 50)  # 限制最大宽度为50
        ws.column_dimensions[column_letter].width = adjusted_width

# =============== 主流程（仅 4 个入参） ===============
def extract_tumor_volume_for_word(xlsx_path: str,
                                  out_xlsx_path: str,
                                  end_day: int) -> bool:
    """
    读取"实验动物荷瘤体积(mm3)"与实验设计，导出 7-2 表：
    组别 | 受试品 | 给药前（均数±SD） | 第{end_day}天（均数±SD） | TGITV(%) | p | 肿瘤清除比例
    """
    C = CONFIG
    control_group = "G1"  # 对照组固定为G1
    try:
        wb = load_workbook(xlsx_path, data_only=True)

        # 数据页
        sheet_data_name = find_existing_sheet(wb, C["SHEET_DATA"])
        if not sheet_data_name:
            raise RuntimeError(f"未找到数据页，候选：{C['SHEET_DATA']}")
        ws: Worksheet = wb[sheet_data_name]

        # 1) 锚点：找包含"实验动物荷瘤体积/ Tumor Volume"等关键词的单元格
        anchor = None
        for r in range(1, ws.max_row + 1):
            hit = False
            for c in range(1, ws.max_column + 1):
                t = val_eff(ws, r, c)
                if contains_any(t, C["ANCHOR_CONTAINS"]):
                    anchor = (r, c)
                    hit = True
                    break
            if hit:
                break
        if not anchor:
            raise RuntimeError(f"未找到包含锚点关键词的单元格，候选：{C['ANCHOR_CONTAINS']}")
        r0, c0 = anchor

        # 两个关键列
        group_col = c0        # A列：G1/G2...
        stat_col  = c0 + 1    # B列：均数/标准误/TGITV/CV值

        # 2) 结束列：窗口 r0..r0+N，找首个"该列全空"→ 前一列为结束列
        r_end_window = min(ws.max_row, r0 + C["DAYS_ROW_LOOKAHEAD"])
        end_col = None
        for c in range(c0, ws.max_column + 1):
            if all(is_empty(ws, r, c) for r in range(r0, r_end_window + 1)):
                end_col = c - 1
                break
        if end_col is None or end_col < c0:
            raise RuntimeError("未能确定结束列。")

        # 3) 结束行：限定 [c0..end_col]，自 r0 向下找首个"整行全空"，上一行即 end_row
        end_row = None
        for r in range(r0, ws.max_row + 1):
            if all(is_empty(ws, r, c) for c in range(c0, end_col + 1)):
                end_row = r - 1
                break
        if end_row is None or end_row < r0:
            # 如果找不到空行，直接取工作表的最后一行作为结束行
            end_row = ws.max_row

        # 4) 找"分组后天数"→ 下一行是天数行 → 找 0 与 end_day 列
        rA = cA = None
        for rr in range(r0, end_row + 1):
            for cc in range(c0, end_col + 1):
                if contains_any(val_eff(ws, rr, cc), C["DAYS_HEADER"]):
                    rA, cA = rr, cc
                    break
            if rA is not None:
                break
        if rA is None:
            raise RuntimeError(f"未在表格矩形内找到天数表头，候选：{C['DAYS_HEADER']}")
        r_days = rA + 1

        day_to_col = {}
        for cc in range(cA, end_col + 1):
            txt = norm(val_eff(ws, r_days, cc))
            m = re.match(r"^\D*(-?\d+)\D*$", txt)
            if m:
                day_to_col[int(m.group(1))] = cc
        if 0 not in day_to_col or end_day not in day_to_col:
            raise RuntimeError(f"天数行未找到 0 或 {end_day}。识别到: {sorted(day_to_col)}")
        col0, colN = day_to_col[0], day_to_col[end_day]

        # 5) 分组：用 group_col 找 Gx；组块结束 = 下个起点 - 1
        patG = re.compile(C["GROUP_PATTERN"], re.IGNORECASE)
        group_starts = [r for r in range(r0, end_row + 1) if patG.match(norm(val_eff(ws, r, group_col)))]
        if not group_starts:
            raise RuntimeError("未找到任何组别（G1/G2/...）。")

        rows_out = []
        # 保存长表用于 Dunnett
        long_rows = []
        per_group_values = {}

        for i, rs in enumerate(group_starts):
            re_ = group_starts[i + 1] - 1 if i < len(group_starts) - 1 else end_row
            group_name = norm(val_eff(ws, rs, group_col))

            # 在 stat_col（B列）里找 "均数/标准误/TGITV"
            r_mean = r_sd = r_tgi = None
            for rr in range(rs, re_ + 1):
                t = norm(val_eff(ws, rr, stat_col))
                if (r_mean is None) and contains_any(t, C["MEAN_LABELS"]):
                    r_mean = rr
                if (r_sd is None) and (contains_any(t, C["SD_LABELS"]) or t.upper() == "SD"):
                    r_sd = rr
                if (r_tgi is None) and contains_any(t, C["TGITV_LABELS"]):
                    r_tgi = rr
                if r_mean and r_sd and r_tgi:
                    break

            # 读取统计值
            m0 = mN = s0 = sN = None
            if r_mean:
                m0 = parse_float(val_eff(ws, r_mean, col0))
                mN = parse_float(val_eff(ws, r_mean, colN))
            if r_sd:
                s0 = parse_float(val_eff(ws, r_sd, col0))
                sN = parse_float(val_eff(ws, r_sd, colN))

            # TGITV（对照组通常空白；若源表无该行则留空）
            tgi = None
            if r_tgi:
                # 获取单元格的值，直接处理数值类型
                cell_value = ws.cell(row=r_tgi, column=colN).value
                if cell_value is not None:
                    try:
                        # 直接将数值乘以100并保留一位小数
                        tgi = f"{round(float(cell_value) * 100, 1)}"
                    except (ValueError, TypeError):
                        tgi = None

            # 个体原始值（用于 Dunnett）：从"组别行 rs"到"均数行-1"都视作动物行
            if r_mean:
                end_anim_row = r_mean - 1
            else:
                end_anim_row = re_
            values = []
            for rr in range(rs, end_anim_row + 1):  # 兼容源表把个体值写在组别同一行的情形
                v = parse_float(val_eff(ws, rr, colN))
                if v is not None:
                    try:
                        # 确保P值计算使用取整数后的值
                        v = round(float(v))  # 取整数
                        values.append(v)
                        long_rows.append({"group": group_name, "volume": float(v)})
                    except (ValueError, TypeError):
                        continue
            per_group_values[group_name] = values

            rows_out.append({
                "组别": group_name,
                "分组天均值": fmt_pm(m0, s0, 0),  # 0表示只保留整数
                "结束天均值": fmt_pm(mN, sN, 0),  # 0表示只保留整数
                "TGITV": (tgi if (tgi is not None and group_name != control_group) else ""),
            })

        df = pd.DataFrame(rows_out)

        # 6) 实验设计：组别→受试品(剂量)
        sheet_design_name = find_existing_sheet(wb, C["SHEET_DESIGN"])
        if sheet_design_name:
            wsD: Worksheet = wb[sheet_design_name]
            hdr = None
            for r in range(1, wsD.max_row + 1):
                row_txt = " ".join(norm(val_eff(wsD, r, c)) for c in range(1, wsD.max_column + 1))
                if contains_any(row_txt, C["DESIGN_GROUP_HEADER"]) and \
                   contains_any(row_txt, C["DESIGN_DRUG_HEADERS"]) and \
                   contains_any(row_txt, C["DESIGN_DOSE_HEADERS"]):
                    hdr = r
                    break
            if hdr is None:
                raise RuntimeError("实验设计页未找到表头。")

            col_g = col_d = col_do = None
            for c in range(1, wsD.max_column + 1):
                t = norm(val_eff(wsD, hdr, c))
                if (col_g is None) and contains_any(t, C["DESIGN_GROUP_HEADER"]):
                    col_g = c
                if (col_d is None) and contains_any(t, C["DESIGN_DRUG_HEADERS"]):
                    col_d = c
                if (col_do is None) and contains_any(t, C["DESIGN_DOSE_HEADERS"]):
                    col_do = c

            mapping = {}
            blank = 0
            for r in range(hdr + 1, wsD.max_row + 1):
                g = norm(val_eff(wsD, r, col_g)) if col_g else ""
                if g == "":
                    blank += 1
                    if blank >= 2:
                        break
                    continue
                blank = 0
                drug = norm(val_eff(wsD, r, col_d)) if col_d else ""
                dose = norm(val_eff(wsD, r, col_do)) if col_do else ""
                combo = f"{drug}({dose})" if (drug and dose) else (drug or (f"({dose})" if dose else ""))
                if combo:
                    mapping.setdefault(g, []).append(combo)

            if mapping:
                df_map = pd.DataFrame([{"组别": k, "受试品": ", ".join(v)} for k, v in mapping.items()])
                df = df.merge(df_map, on="组别", how="left")
                df.insert(1, "受试品", df.pop("受试品"))
            else:
                df.insert(1, "受试品", "")

        # 7) P 值（Dunnett，对照组默认 control_group）
        from P_compute import calculate_dunnett_json
        import json

        if long_rows:
            dunnett_res = calculate_dunnett_json(json.dumps(long_rows, ensure_ascii=False),
                                                 control=control_group)
            sp_map = {}
            for r in dunnett_res:
                g = (r.get("group") or "").strip()
                stars = (r.get("Summary") or "").strip()     # '**' 或 '' 或 'ns'
                pval  = (r.get("P-Value") or "").strip()     # '0.0056' 等
                sp_map[g] = pval if stars == "ns" else f"{stars}{pval}"
            df["P值"] = df["组别"].map(lambda g: "" if g == control_group else sp_map.get(g, ""))
        else:
            df["P值"] = ""

        # 8) 列顺 & 肿瘤清除比例（空列）
        wanted = ["组别", "受试品", "分组天均值", "结束天均值", "TGITV", "P值"]
        df = df[[c for c in wanted if c in df.columns]]
        df.insert(df.columns.get_loc("P值") + 1, "肿瘤清除比例", "")

        # 9) 空值标准化
        df = df.fillna("-").replace("", "-")

        # 10) 写出到 out_xlsx_path / OUT_SHEET
        import os
        if not os.path.exists(out_xlsx_path):
            # 若输出文件不存在，新建一个
            book = Workbook()
            # openpyxl 新建的默认表名为 "Sheet"，删掉以保持干净
            default_sheet = book.active
            book.remove(default_sheet)
            book.save(out_xlsx_path)

        book = load_workbook(out_xlsx_path)
        if C["OUT_SHEET"] in book.sheetnames:
            book.remove(book[C["OUT_SHEET"]])
        ws_out = book.create_sheet(C["OUT_SHEET"])

        # 先写表头
        for c_idx, col_name in enumerate(df.columns, 1):
            ws_out.cell(row=1, column=c_idx, value=col_name)
        # 再写数据（从第 2 行起）
        for r_idx, row in enumerate(df.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, 1):
                ws_out.cell(row=r_idx, column=c_idx, value=value)

        auto_adjust_column_width(ws_out)
        # 先保存工作簿，然后再添加GraphPad工作表
        book.save(out_xlsx_path)
        
        # 添加GraphPad工作表
        if add_graphpad_sheet(out_xlsx_path, "7-2实验动物荷瘤体积数据", per_group_values):
            print(f"成功添加GraphPad使用工作表")
        else:
            print(f"添加GraphPad使用工作表失败")

        print(f"OK: 生成 {out_xlsx_path} / {C['OUT_SHEET']}，{len(df)} 行")
        return True

    except Exception as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        return False

# =============== 入口（3 参数） ===============
if __name__ == "__main__":
    input_path = "D:/TianBa_AI/Code/docs/temp/project_report/25P080002_Final.xlsx"
    output_path = "D:/TianBa_AI/Code/docs/temp/project_report/25P080002_明细.xlsx"
    end_day = 14
    
    ok = extract_tumor_volume_for_word(input_path, output_path, end_day)
    sys.exit(0 if ok else 1)
