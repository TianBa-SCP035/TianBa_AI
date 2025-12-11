# -*- coding: utf-8 -*-
import re
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from graphpad_helper import add_graphpad_sheet


# =============== 变量区（所有可调参数都在这里） ===============
CONFIG = {
    # 工作表与输出
    "SHEET_WEIGHT": ["实验数据汇总", "Study Data"],   # 体重数据页
    "SHEET_DESIGN": ["实验设计", "Study Design"],     # 实验设计页（可不存在）
    "OUT_SHEET": "form_7_1",                         # 输出工作表名称

    # 关键文本/模式（中文中可能出现全角空格或括号，后续会正则规整）
    "ANCHOR_CONTAINS": ["实验动物体重克", "Animal Weight（g）"],   # 锚点关键词
    "DAYS_HEADER": ["分组后天数", "Days Post Grouping"],        # 天数行的表头关键词

    # 统计行识别（位于组别块内第二列）
    "MEAN_LABELS": ["均数", "Average"],
    "SD_LABELS": ["标准误", "Standard Error of the Mean"],

    # 设计页表头（允许出现任意一个候选）
    "DESIGN_GROUP_HEADER": ["组别", "Groups"],
    "DESIGN_DRUG_HEADERS": ["处理方式", "Treatment"],
    "DESIGN_DOSE_HEADERS": ["剂量", "Dosages"],

    # 其它启发式
    "GROUP_PATTERN": r"^\s*G\d+\b",   # 组别识别（首列形如 G1/G2/...）
    "DAYS_ROW_LOOKAHEAD": 5,          # 在锚点下方多少行窗口内判定"结束列"
    "DECIMALS": 1,                    # 均值/差值保留小数位
}


# =============== 基础工具 ===============
def norm(s) -> str:
    if s is None:
        return ""
    s = str(s).replace("\u3000", " ").strip()
    return re.sub(r"\s+", " ", s)

def contains_any(text, patterns) -> bool:
    """检查文本是否包含任意一个模式字符串;patterns 可为 str 或 list"""
    if not text or not patterns:
        return False
    if isinstance(patterns, str):
        patterns = [patterns]
    text_norm = norm(text)
    # 对文本也应用与文件相同的处理：移除括号和空格
    text_norm = re.sub(r"[（）()\s]", "", text_norm)
    for pattern in patterns:
        # 对关键词也应用相同的处理
        pattern_norm = norm(pattern)
        pattern_norm = re.sub(r"[（）()\s]", "", pattern_norm)
        if pattern_norm in text_norm:
            return True
    return False

def find_existing_sheet(wb, sheet_names) -> str:
    """从候选工作表名称中返回第一个存在的名称；否则返回空字符串"""
    if isinstance(sheet_names, str):
        sheet_names = [sheet_names]
    for name in sheet_names or []:
        if name in wb.sheetnames:
            return name
    return ""

def in_merged(ws: Worksheet, r: int, c: int):
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:
            return rng
    return None

def val_eff(ws: Worksheet, r: int, c: int):
    v = ws.cell(row=r, column=c).value
    if v is not None:
        return v
    rng = in_merged(ws, r, c)
    if rng:
        return ws.cell(rng.min_row, rng.min_col).value
    return None

def is_empty(ws: Worksheet, r: int, c: int) -> bool:
    return norm(val_eff(ws, r, c)) == ""

def parse_float(x):
    s = norm(x)
    if s in ("", "-", "—", "–"):
        return None
    s = s.replace(",", "").replace("%", "")
    try:
        return float(s)
    except Exception:
        m = re.search(r"(-?\d+(?:\.\d+)?)", s)
        return float(m.group(1)) if m else None

def fmt_pm(m, sd, d=1):
    if m is None or sd is None:
        return ""
    return f"{m:.{d}f}±{sd:.{d}f}"

def fmt_signed(x, d=1):
    if x is None:
        return ""
    s = f"{x:.{d}f}"
    # 在加减号后都添加空格，保持格式一致
    return f"- {s[1:]}" if s.startswith("-") else f"+ {s}"

def auto_adjust_column_width(ws):
    """自动调整工作表列宽以适应内容"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 6, 50)  # 限制最大宽度为50
        ws.column_dimensions[column_letter].width = adjusted_width


# =============== 主流程（仅 3 个入参） ===============
def extract_weight_for_word(xlsx_path: str, out_xlsx_path: str, end_day: int) -> bool:
    """
    读取体重汇总与实验设计，导出整理表。
    入参：数据表格路径、写入表格路径、结束天数。
    返回:True 表示成功,False 表示失败。
    """
    C = CONFIG

    try:
        wb = load_workbook(xlsx_path, data_only=True)

        # 体重数据页
        sheet_weight_name = find_existing_sheet(wb, C["SHEET_WEIGHT"])
        if not sheet_weight_name:
            raise RuntimeError(f"未找到体重数据页，候选：{C['SHEET_WEIGHT']}")
        ws: Worksheet = wb[sheet_weight_name]

# ========================= 第一部分：均数标准误的拼接 ==================================
        # 1) 锚点：从(1,1)起，首个包含锚点关键词的单元格
        anchor = None
        for r in range(1, ws.max_row + 1):
            hit = False
            for c in range(1, ws.max_column + 1):
                t = norm(val_eff(ws, r, c))
                # 移除了额外的括号和空格处理，因为contains_any函数内部已经处理了
                if contains_any(t, C["ANCHOR_CONTAINS"]):
                    anchor = (r, c)
                    hit = True
                    break
            if hit:
                break
        if not anchor:
            raise RuntimeError(f"未找到包含锚点关键词的单元格，候选：{C['ANCHOR_CONTAINS']}")
        r0, c0 = anchor

        # 两个关键列
        group_col = c0        # A列：G1/G2...
        stat_col  = c0 + 1    # B列：均数/标准误/CV值

        # 2) 结束列：窗口 r0..r0+N，找首个"该列全空"→ 前一列为结束列
        r_end_window = min(ws.max_row, r0 + C["DAYS_ROW_LOOKAHEAD"])
        end_col = None
        for c in range(c0, ws.max_column + 1):
            if all(is_empty(ws, r, c) for r in range(r0, r_end_window + 1)):
                end_col = c - 1
                break
        if end_col is None or end_col < c0:
            raise RuntimeError("未能确定结束列。")

        # 3) 结束行：限定 [c0..end_col]，自 r0 向下找首个"整行全空"，上一行即 end_row
        end_row = None
        for r in range(r0, ws.max_row + 1):
            if all(is_empty(ws, r, c) for c in range(c0, end_col + 1)):
                end_row = r - 1
                break
        if end_row is None or end_row < r0:
            raise RuntimeError("未能确定结束行。")

        # 4) 找"分组后天数"→ 下一行是天数行 → 找 0 与 end_day 列
        rA = cA = None
        for rr in range(r0, end_row + 1):
            for cc in range(c0, end_col + 1):
                if contains_any(val_eff(ws, rr, cc), C["DAYS_HEADER"]):
                    rA, cA = rr, cc
                    break
            if rA is not None:
                break
        if rA is None:
            raise RuntimeError(f"未在表格矩形内找到天数表头，候选：{C['DAYS_HEADER']}")
        r_days = rA + 1

        day_to_col = {}
        for cc in range(cA, end_col + 1):
            txt = norm(val_eff(ws, r_days, cc))
            m = re.match(r"^\D*(-?\d+)\D*$", txt)
            if m:
                day_to_col[int(m.group(1))] = cc
        if 0 not in day_to_col or end_day not in day_to_col:
            raise RuntimeError(f"天数行未找到 0 或 {end_day}。识别到: {sorted(day_to_col)}")
        col0, colN = day_to_col[0], day_to_col[end_day]

        # 5) 分组：用 group_col 找 Gx；组块结束 = 下个起点 - 1
        patG = re.compile(C["GROUP_PATTERN"], re.IGNORECASE)
        group_starts = [r for r in range(r0, end_row + 1) if patG.match(norm(val_eff(ws, r, group_col)))]
        if not group_starts:
            raise RuntimeError("未找到任何组别（G1/G2/...）。")

        rows_out = []
        for i, rs in enumerate(group_starts):
            re_ = group_starts[i + 1] - 1 if i < len(group_starts) - 1 else end_row
            group_name = norm(val_eff(ws, rs, group_col))

            # 在 stat_col（B列）里找"均数/标准误"
            r_mean = r_sd = None
            for rr in range(rs, re_ + 1):
                t = norm(val_eff(ws, rr, stat_col))
                if (r_mean is None) and contains_any(t, C["MEAN_LABELS"]):
                    r_mean = rr
                if (r_sd is None) and (contains_any(t, C["SD_LABELS"]) or t.upper() == "SD"):
                    r_sd = rr
                if r_mean and r_sd:
                    break

            m0 = mN = s0 = sN = None
            if r_mean:
                m0 = parse_float(val_eff(ws, r_mean, col0))
                mN = parse_float(val_eff(ws, r_mean, colN))
            if r_sd:
                s0 = parse_float(val_eff(ws, r_sd, col0))
                sN = parse_float(val_eff(ws, r_sd, colN))
            
            # 添加容差值处理浮点数精度问题，然后直接使用round函数
            m0_rounded = round(m0 + 1e-06, C["DECIMALS"]) if m0 is not None else None
            mN_rounded = round(mN + 1e-06, C["DECIMALS"]) if mN is not None else None
            delta = (mN_rounded - m0_rounded) if (mN_rounded is not None and m0_rounded is not None) else None

            rows_out.append({
                "组别": group_name,
                "分组天均值": fmt_pm(m0_rounded, s0, C["DECIMALS"]),
                "结束天均值": fmt_pm(mN_rounded, sN, C["DECIMALS"]),
                "差值": fmt_signed(delta, C["DECIMALS"]),
            })

        df = pd.DataFrame(rows_out)

# ============================= 第二部分：受试品拼接 ================================
        # 6) 实验设计：组别→受试品(剂量)
        sheet_design_name = find_existing_sheet(wb, C["SHEET_DESIGN"])
        if sheet_design_name:
            wsD: Worksheet = wb[sheet_design_name]
            hdr = None
            for r in range(1, wsD.max_row + 1):
                row_txt = " ".join(norm(val_eff(wsD, r, c)) for c in range(1, wsD.max_column + 1))
                if contains_any(row_txt, C["DESIGN_GROUP_HEADER"]) and \
                   contains_any(row_txt, C["DESIGN_DRUG_HEADERS"]) and \
                   contains_any(row_txt, C["DESIGN_DOSE_HEADERS"]):
                    hdr = r
                    break
            if hdr is None:
                raise RuntimeError("实验设计页未找到表头。")

            col_g = col_d = col_do = None
            for c in range(1, wsD.max_column + 1):
                t = norm(val_eff(wsD, hdr, c))
                if (col_g is None) and contains_any(t, C["DESIGN_GROUP_HEADER"]):
                    col_g = c
                if (col_d is None) and contains_any(t, C["DESIGN_DRUG_HEADERS"]):
                    col_d = c
                if (col_do is None) and contains_any(t, C["DESIGN_DOSE_HEADERS"]):
                    col_do = c

            mapping = {}
            blank = 0
            for r in range(hdr + 1, wsD.max_row + 1):
                g = norm(val_eff(wsD, r, col_g)) if col_g else ""
                if g == "":
                    blank += 1
                    if blank >= 2:
                        break
                    continue
                blank = 0
                drug = norm(val_eff(wsD, r, col_d)) if col_d else ""
                dose = norm(val_eff(wsD, r, col_do)) if col_do else ""
                combo = f"{drug}({dose})" if (drug and dose) else (drug or (f"({dose})" if dose else ""))
                if combo:
                    mapping.setdefault(g, []).append(combo)

            if mapping:
                df_map = pd.DataFrame([{"组别": k, "受试品": ", ".join(v)} for k, v in mapping.items()])
                df = df.merge(df_map, on="组别", how="left")
                df.insert(1, "受试品", df.pop("受试品"))
            else:
                df.insert(1, "受试品", "")

# ========================= 第三部分：P值计算（包括组别行） =================================
        from P_compute import calculate_dunnett_json
        import json

        long_rows = []           # 供 puc 的长表
        per_group_values = {}    # 可选：调试查看每组 end-day 原始值

        for i, rs in enumerate(group_starts):
            re_ = group_starts[i + 1] - 1 if i < len(group_starts) - 1 else end_row
            group_name = norm(val_eff(ws, rs, group_col))

            # 找到该组的"均数"行（你前面已实现 r_mean / r_sd 的搜索）
            r_mean = None
            for rr in range(rs, re_ + 1):
                t = norm(val_eff(ws, rr, stat_col))
                if contains_any(t, C["MEAN_LABELS"]):
                    r_mean = rr
                    break

            # 从 组别行 到 "均数行"上一行，全部视为动物行；读取 end-day 列
            values = []
            if r_mean:  # 正常路径
                end_anim_row = r_mean - 1
            else:       # 兜底：若未识别到均数行，则以组块末行为封口
                end_anim_row = re_

            for rr in range(rs, end_anim_row + 1):  # 修改：从 rs 开始，而不是 rs + 1
                v = parse_float(val_eff(ws, rr, colN))
                if v is not None:
                    # 使用Decimal进行精确四舍五入，与Excel保持一致
                    v = round(v + 1e-06, 1)  # 添加容差值，保留1位小数
                    values.append(v)
                    long_rows.append({"group": group_name, "volume": float(v)})
            per_group_values[group_name] = values

        # 打印 per_group_values 以便调试
        print("\n=== 每组 end-day 原始值（包括组别行） ===")
        for group, values in per_group_values.items():
            print(f"{group}: {values}")
        print("=====================================\n")

        # 调用 puc 计算 Dunnett（默认 G1 为对照；如需自定义可从设计页映射）
        if long_rows:
            control_group = "G1"
            dunnett_res = calculate_dunnett_json(json.dumps(long_rows, ensure_ascii=False),
                                             control=control_group)
            # 合并星值与P值，但当星值为"ns"时不拼接
            sp_map = {}
            for r in dunnett_res:
                g = r.get("group")
                stars = (r.get("Summary") or "").strip()     # '**' 或 '' 或 'ns'
                pval  = (r.get("P-Value") or "").strip()     # '0.0056' 等
                # 当星值为"ns"时，只显示P值；否则显示星值+P值
                sp_map[g] = pval if stars == "ns" else f"{stars}{pval}"
            df["P值"] = df["组别"].map(lambda g: "" if g == control_group else sp_map.get(g, ""))
        else:
            df["P值"] = ""
        
        # 将P值列移动到第5列（差值之前）
        if "P值" in df.columns:
            p_values = df.pop("P值")
            df.insert(4, "P值", p_values)
        else:
            # 如果P值列不存在，添加一个空的P值列
            df.insert(4, "P值", "")

# =============================== 结束：全部写出 =====================================
        # 将所有空值（包括NaN和空字符串）替换为"-"
        df = df.fillna("-")  # 替换NaN
        df = df.replace("", "-")  # 替换空字符串
        
        # 简化文件写入逻辑：使用openpyxl直接操作工作簿
        import os
        if not os.path.exists(out_xlsx_path):
            print(f"[ERROR] 输出文件不存在: {out_xlsx_path}")
            return False
            
        # 文件存在，加载现有工作簿
        book = load_workbook(out_xlsx_path)
        # 如果已存在同名工作表，先删除
        if C["OUT_SHEET"] in book.sheetnames:
            book.remove(book[C["OUT_SHEET"]])
        # 创建新工作表
        ws = book.create_sheet(C["OUT_SHEET"])
        # 写入表头
        for c_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=c_idx, value=col_name)
        # 将DataFrame数据写入工作表
        for r_idx, row in enumerate(df.values, 2):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        # 自动调整列宽
        auto_adjust_column_width(ws)
        
        # 保存工作簿（添加GraphPad工作表之前）
        book.save(out_xlsx_path)
        
        # 添加GraphPad使用工作表
        print("[INFO] 正在添加GraphPad使用工作表...")
        if add_graphpad_sheet(out_xlsx_path, "7-1实验动物体重数据", per_group_values):
            print("[SUCCESS] GraphPad使用工作表添加成功")
        else:
            print("[ERROR] GraphPad使用工作表添加失败")

        print(f"OK: 生成 {out_xlsx_path} / {C['OUT_SHEET']}，{len(df)} 行")
        return True

    except Exception as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        return False

# =============== 入口（仅 3 个参数，支持默认值） ===============
if __name__ == "__main__":
    # 直接使用完整路径和固定参数
    input_path = "D:/TianBa_AI/Code/docs/temp/project_report/25P118604_Final.xlsx"
    output_path = "D:/TianBa_AI/Code/docs/temp/project_report/25P118604_明细.xlsx"
    end_day = 20
    
    ok = extract_weight_for_word(input_path, output_path, end_day)
    sys.exit(0 if ok else 1)
