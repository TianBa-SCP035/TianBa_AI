# -*- coding: utf-8 -*-

import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
from translate import Translator

def contains_chinese(text):
    """检查文本是否包含中文字符"""
    return bool(re.search(r'[\u4e00-\u9fff]', text))

def clean_translated_text(text, translator, cache):
    """清理翻译后的文本,移除HTML标签和多余空格，使用缓存避免重复翻译"""
    # 如果缓存中已有翻译结果，直接返回
    if text in cache:
        return text, cache[text], None
    
    try:
        translated = re.sub(r'<[^>]+>', '', translator.translate(text))
        translated = re.sub(r'\s+', ' ', translated).strip()
        # 将翻译结果存入缓存
        cache[text] = translated
        return text, translated, None
    except Exception as e:
        return text, None, str(e)

def translate_excel_region(file_path, sheet_name, start_row, end_row, start_col, end_col, direction=0, max_workers=20):
    """翻译Excel中指定区域的内容"""
    try:
        wb = load_workbook(file_path)
        
        if sheet_name not in wb.sheetnames:
            print(f"错误: 工作表 '{sheet_name}' 不存在。可用工作表: {', '.join(wb.sheetnames)}")
            return False
        ws = wb[sheet_name]
        #设置翻译方向，1为英翻中，其他为中翻英
        from_lang, to_lang = ("en", "zh-cn") if direction == 1 else ("zh-cn", "en")
        direction_text = "英翻中" if direction == 1 else "中翻英"
        
        # 收集需要翻译的单元格
        cells_to_translate = []
        for row in range(start_row, end_row + 1):
            for col in range(ord(start_col.upper()) - ord('A') + 1, ord(end_col.upper()) - ord('A') + 2):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    # 只翻译包含中文字符的单元格（中翻英时）
                    # 或者不包含中文字符的单元格（英翻中时）
                    text = cell.value.strip()
                    if (direction == 1 and not contains_chinese(text)) or (direction != 1 and contains_chinese(text)):
                        cells_to_translate.append((cell, text))
        
        if not cells_to_translate:
            print("没有找到需要翻译的文本内容")
            return False
            
        print(f"开始翻译工作表 '{sheet_name}' 中的 {len(cells_to_translate)} 个单元格 ({direction_text})")
        
        # 并发翻译
        translator = Translator(from_lang=from_lang, to_lang=to_lang)
        translation_cache = {}
        completed_count = 0
        
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_cell = {
                executor.submit(clean_translated_text, text, translator, translation_cache): (cell, text)
                for cell, text in cells_to_translate
            }
            
            for future in as_completed(future_to_cell):
                cell, original_text = future_to_cell[future]
                original, translated, error = future.result()
                
                if error:
                    print(f"翻译失败: 行{cell.row}, 列{chr(cell.column+64)}, 错误: {error}")
                else:
                    cell.value = translated
                    completed_count += 1
                    print(f"翻译进度: {completed_count}/{len(cells_to_translate)} ({completed_count/len(cells_to_translate)*100:.1f}%)", end='\r')
        
        wb.save(file_path)
        print(f"\n翻译完成，已保存到: {file_path}")
        return True
    except Exception as e:
        print(f"翻译Excel时出错: {str(e)}")
        return False

if __name__ == "__main__":
    translate_excel_region("25P089601_明细.xlsx", "明细", 1, 40, "B", "B")